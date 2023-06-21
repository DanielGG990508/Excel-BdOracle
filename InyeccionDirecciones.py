# -*- coding: utf-8 -*-
"""
Created on Mon May 29 12:45:35 2023

@author: 10042891
"""
import cx_Oracle
import openpyxl
import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
# Se establece la conexión a la base de datos
conn = cx_Oracle.connect('RVICEPRE/0rAcleDevVP2@10.204.14.120:1521')
# Definir la ruta y el nombre del archivo excel
ruta_excel = "C:/Users/10042891/.spyder-py3/DatosRed/PrimerCircuito/InyeccionDireciones.xlsx"

ruta_excel1 = "C:/Users/10042891/.spyder-py3/DatosRed/PrimerCircuito/InyeccionDirecionesFaltantes.xlsx"
# Cargar el archivo excel
workbook = openpyxl.load_workbook(ruta_excel)
# Se crea un cursor
cursor = conn.cursor()

# Se define la consulta de inserción
sqlInsert = """INSERT INTO RVICEPRE.TAVPDIRECCIONPT2 ( FIIDPERFK, FCCALLE, FINOEXT, FINOINT, FCREFERENCIA, FCTIPO, FCUSERACT, FDFECTHACT, FIIDCOLFK, FIIDCDFK, FIIDEDO) 
               VALUES (:1, :2, :3, :4, :5, :6, :7, :8, :9, :10, :11)"""
        
sqlConsulta = "SELECT * FROM rvicepre.tcvpccoloniat2 WHERE fiidedofk=:param1 AND lower(fccolonia) LIKE lower(:param2)"

sqlConsCont = "select fiidotrapersona from rvicepre.tavpcontacto where fiidpersonafk = :idPersonaP"
lista_idsCon = []
worksheet = workbook["InyeccionDireciones"]
iterador = 0
idsPersona = 26
fecha_actual = datetime.datetime.now().strftime("%d/%m/%y")
idEstado=9
idColonia = 0
idCiudad = 0
# Crear una lista para almacenar las filas sin coincidencia
filas_sin_coincidencia = []

fila_actual = 2  # Número de la primera fila de datos en la hoja de cálculo

for row in worksheet.iter_rows(min_row=3, values_only=True):
    calle = row[1]
    nomExt = row[2]
    colonia = str(row[3]).strip()
########seccion para insertar direcciones a elementos de la red 
    cursor.execute(sqlConsCont, {'idPersonaP': idsPersona})
    resContacto = cursor.fetchall()
    for resultado in resContacto:
        lista_idsCon.append(resultado[0]) 
    print(lista_idsCon)
########################################################################
    print(f'esta es la colonia {colonia}')
    cursor.execute(sqlConsulta, {'param1': 9, 'param2': '%' + colonia + '%'})
    resultados = cursor.fetchone()
    
    if resultados is not None:
        # Obtener los valores de los atributos deseados
        idColonia = resultados[0]  # Primer atributo
        idCiudad = resultados[5]  # Segundo atributo
        
    else:
        # Almacenar el número de fila sin coincidencia en la lista
        filas_sin_coincidencia.append(fila_actual)
    print(iterador)
    iterador+=1
    fila_actual += 1  # Incrementar el número de fila actual
###este for es inecesario si son personas solo es funcional para red
    for i in lista_idsCon:
        cursor.execute(sqlInsert, [i , calle, nomExt, 0, "Sin Referencia", "Trabajo", "DanielDireccionRed", fecha_actual, idColonia, idCiudad, idEstado])
    #cursor.execute(sqlInsert, [idsPersona , calle, nomExt, 0, "Sin Referencia", "Trabajo", "DanielDireccionRed", fecha_actual, idColonia, idCiudad, idEstado])
    idsPersona+=1
    lista_idsCon = []
# Aplicar el formato a las filas sin coincidencia
no_coincidencia_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
for num_fila in filas_sin_coincidencia:
    for i, valor_celda in enumerate(worksheet[num_fila]):
        # Obtener la referencia de la celda correspondiente
        columna = get_column_letter(i + 1)
        celda = worksheet[columna + str(num_fila)]
        celda.fill = no_coincidencia_fill


# Guardar el archivo con el nuevo formato
ruta_nuevo_excel = "C:/Users/10042891/.spyder-py3/DatosRed/PrimerCircuito/DireccionesFaltantesContactos.xlsx"
workbook.save(ruta_nuevo_excel)

conn.commit()
cursor.close()
conn.close()
print("conexion exitosa SE INSERTARON :",{iterador},"datos")

# Se cierra el cursor y la conexión a la base de datos
